#! python
# duesReminder.py - script sends emails based on payment status in .xlsx file
# usage: python duesreminder.py <your_mail> <your_password>
# !! Working only with correct mail and password!

import openpyxl, smtplib, sys

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']
last_col = sheet.max_column
latest_month = sheet.cell(row=1, column=last_col).value

unpaid_members = {}

if len(sys.argv) != 3:
    print('USAGE: python duesreminder.py <your_mail> <your_password> ')
    exit(-1)

for row_num in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=row_num, column=last_col).value
    if payment != 'paid':
        name = sheet.cell(row=row_num, column = 1).value
        email = name = sheet.cell(row=row_num, column = 2).value
        unpaid_members[name] = email

smtp_obj = smtplib.SMTP('smtp.example.com', 587)
smtp_obj.ehlo()
smtp_obj.starttls()
try:
    smtp_obj.login(sys.argv[1], sys.argv[2])
except:
    print('Sorry, have difficulty with access to your email')
    exit(-1)

for name, email in unpaid_members.items():
    body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid money for %s. Plese make this payment as soon as possible. Thankt you" % (latest_month, name, latest_month)
    print(f'Sending email to {email}...')
    send_mail_status = smtp_obj.sendmail('examplemailaddress@example.com', email, body)

    if send_mail_status != {}:
        print(f'There was a problem sending email to {email}: {send_mail_status}')
smtp_obj.quit()